import openpyxl
from collections import Counter

# 最新の蓄積ファイルを指定
filename = "google_research_history.xlsx"

# Excelファイルを読み込む
wb = openpyxl.load_workbook(filename)
ws = wb.active

# キーワードが入っている列（B列＝2列目）を指定
keyword_column_index = 2

keywords = []

# 2行目から最終行までループしてキーワードを回収
for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=keyword_column_index).value
    if val:
        keywords.append(val.strip())

# Counterを使って一瞬で集計
keyword_counts = Counter(keywords)
total_count = len(keywords)

print("\n📊 【最新ニュースデータ 視覚的集計結果】 📊")
print(f"分析した総ユニーク記事数: {total_count} 件")
print("-" * 55)

# 件数が多い順にランキング表示（簡易グラフ付き）
for rank, (word, count) in enumerate(keyword_counts.most_common(), 1):
    percentage = (count / total_count) * 100
    
    # パーセンテージに応じた長さのメーターを作成（2%につき「■」1個、最大50文字分）
    bar_length = int(percentage / 2)
    bar = "■" * bar_length
    
    # 綺麗に整列して表示
    print(f"第 {rank} 位: {word:<5} -> {count:>3} 件 ({percentage:>4.1f}%) {bar}")

print("-" * 55)