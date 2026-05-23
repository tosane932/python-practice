import openpyxl

# 1. 正しいExcelファイル名を設定
input_file = "google_research_history.xlsx"   # 元の1251件のファイル
output_file = "google_research_history_v2.xlsx" # 整理後の綺麗なファイル

# 既存のExcelファイルを読み込む
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# 新しいExcelファイル（整理後用）を作成する
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.title = "整理済みニュース"

# 見出し（ヘッダー）をコピーする（1行目）
header = [cell.value for cell in ws[1]]
new_ws.append(header)

# 一度見つかったURLを記憶しておくためのセット
seen_urls = set()

# カウンター
duplicate_count = 0
unique_count = 0

print(f"「{input_file}」の重複チェックを開始します（E列のURLを基準にします）...")

# URLが「E列（5列目）」にあるので 5 を指定
url_column_index = 5 

# 2行目から最終行までループ処理
for row_idx in range(2, ws.max_row + 1):
    # 行の全セルの値を取得
    row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    
    # URLを取得（インデックスは0から始まるので column_index - 1）
    url = row_values[url_column_index - 1]
    
    # URLが空でなく、まだ見つかっていない場合（新しい記事）
    if url and url not in seen_urls:
        seen_urls.add(url)                  # 見つかったURLリストに登録
        new_ws.append(row_values)           # 新しいシートにレコードを追加
        unique_count += 1
    else:
        # すでに登場したURLの場合（重複記事）
        duplicate_count += 1

# 新しいファイルに保存
new_wb.save(output_file)

print("\n--- 整理結果 ---")
print(f"スキャンした総行数: {ws.max_row - 1} 件")
print(f"削除された重複記事: {duplicate_count} 件")
print(f"残ったユニークな記事: {unique_count} 件")
print(f"整理されたファイルを '{output_file}' として保存しました！")